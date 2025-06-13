from flask import Flask, render_template, request, redirect
import openpyxl
import os

app = Flask(__name__)

# Route for the Register page (displays the form)
@app.route('/register')
def register():
    return render_template('register.html')

# Route for Payment Page (displays the payment form)
@app.route('/payment')
def payment():
    return render_template('payment.html')

# Route to handle form submission (saving data to payment Excel sheet)
@app.route('/payment-submit', methods=['POST'])
def payment_submit():
    if request.method == 'POST':
        name = request.form.get('name')
        amount = request.form.get('amount')

        # Create or open payment Excel file
        if not os.path.exists('payment_data.xlsx'):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Amount"])  # Headers
        else:
            workbook = openpyxl.load_workbook('payment_data.xlsx')
            sheet = workbook.active

        # Append payment data
        sheet.append([name, amount])
        workbook.save('payment_data.xlsx')

        # Redirect to the success page
        return redirect('/payment-success')

# Route for Payment Success Page
@app.route('/payment-success')
def payment_success():
    return render_template('payment-success.html')

# Route to handle registration form submission (saving data to registration Excel sheet)
@app.route('/submit', methods=['POST'])
def submit():
    if request.method == 'POST':
        event = request.form.get('event')
        name = request.form.get('name')
        email = request.form.get('email')
        contact = request.form.get('contact')
        attendees = request.form.get('attendees')

        # Create or open registration Excel file
        if not os.path.exists('data.xlsx'):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Event", "Name", "Email", "Contact", "Number of Attendees"])  # Headers
        else:
            workbook = openpyxl.load_workbook('data.xlsx')
            sheet = workbook.active

        # Append registration data
        sheet.append([event, name, email, contact, attendees])
        workbook.save('data.xlsx')

        # Redirect to the payment page after registration
        return redirect('/payment')

# ✅ Route for Volunteer Page (shows volunteer form)
@app.route('/volunteer')
def volunteer():
    return render_template('volunteer.html')

# ✅ Route to handle volunteer form submission (saving to volunteer Excel sheet)
@app.route('/volunteer-submit', methods=['POST'])
def volunteer_submit():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        interests = request.form.get('interests')

        # Create or open volunteer Excel file
        if not os.path.exists('volunteer_data.xlsx'):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Email", "Phone", "Areas of Interest"])  # Headers
        else:
            workbook = openpyxl.load_workbook('volunteer_data.xlsx')
            sheet = workbook.active

        # Append volunteer data
        sheet.append([name, email, phone, interests])
        workbook.save('volunteer_data.xlsx')

        # Redirect back to volunteer page after submitting
        return redirect('/volunteer')

# Run the app
if __name__ == '__main__':
    app.run(debug=True)
